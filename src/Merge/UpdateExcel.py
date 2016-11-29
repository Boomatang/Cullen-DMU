from datetime import date, datetime
from pprint import pprint
from sys import stdout
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

__author__ = 'Desktop'


def check_revision_dates(tsd_file, main_file):

    tsd_dates = tsd_file.revision

    tsd_key = list(tsd_dates.keys())
    tsd_key = tsd_key[0]
    tsd_date = tsd_dates[tsd_key]
    for file in main_file:
        if tsd_key in file.revision.keys():
            if tsd_date == file.revision[tsd_key]:
                pass
            elif tsd_date > file.revision[tsd_key]:
                file.revision[tsd_key] = tsd_date
            else:
                # FIXME this needs to be handled
                stdout('error please fix me')


def update_flags(file, sheet, row):
    # file = MainFile()

    if file.is_TSD:
        value = 'X'
    else:
        value = ''
    sheet.cell(row=row, column=1).value = value

    if file.is_in_progress:
        value = 'X'
    else:
        value = ''
    sheet.cell(row=row, column=2).value = value

    if file.is_error:
        value = 'X'
    else:
        value = ''
    sheet.cell(row=row, column=3).value = value

    if file.is_ignore:
        value = 'X'
    else:
        value = ''
    sheet.cell(row=row, column=4).value = value


def write_main_data(file, sheet, row):
    # file = MainFile()

    sheet.cell(row=row, column=5).value = file.drawing_number
    sheet.cell(row=row, column=6).value = file.drawing_title

    return row


def update_revision(file, sheet, date_dict, row):
    # file = MainFile()

    keys = file.revision.keys()

    for key in keys:
        if file.revision[key] is None or key is None:
            sheet.cell(row=row, column=3).value = "X"
        else:
            sheet.cell(row=row, column=column_index_from_string(date_dict[file.revision[key]])).value = str(key)


def write_to_register(final_doc, main_files, date_dict):
    wb = openpyxl.load_workbook(final_doc)
    sheet = wb.get_sheet_by_name('Register')
    row = 5

    for file in main_files:
        write_main_data(file, sheet, row)
        update_flags(file, sheet, row)
        update_revision(file, sheet, date_dict, row)
        row += 1
    wb.save(final_doc)


def run(final_doc, source_doc):
    existing_drawings = []
    exist_dates, main_files = main_source(final_doc)
    print('Got dates and main files')
    dates = date_range_from_source(source_doc)
    for i in exist_dates:
        dates.append(i)
    dates = list(set(dates))
    dates.sort()
    main_files.sort()
    dates = add_revision_dates(final_doc, dates)
    print('Added new revisions dates to the sheet')
    tsd_files = get_tsd_files(source_doc)

    for item in main_files:
        existing_drawings.append(item.drawing_number)

    for tsd in tsd_files:
        if tsd.drawing_number not in existing_drawings:
            main_files.append(tsd)
        else:
            check_revision_dates(tsd, main_files)

    write_to_register(final_doc, main_files, dates)


    print('I ran')


def get_tsd_files(doc):
    wb = openpyxl.load_workbook(doc)
    sheet = wb.get_sheet_by_name('Sheet1')
    tsd_files = []
    max_rows = sheet.max_row
    row = 2

    while row <= max_rows:
        # print('Working on row : {}'.format(row))
        x = MainFile()
        x.build_tsd_file(row, sheet)
        tsd_files.append(x)
        row += 1

    return tsd_files


def add_revision_dates(doc, rev_dates):
    wb = openpyxl.load_workbook(doc)
    sheet = wb.get_sheet_by_name('Register')

    day_number = 1
    month_number = 2
    year_number = 3
    start_col = 8
    dates_dict = {}

    for dates in rev_dates:

        sheet.cell(row=day_number, column=start_col).value = dates.day
        sheet.cell(row=month_number, column=start_col).value = dates.month
        sheet.cell(row=year_number, column=start_col).value = dates.year

        dates_dict.setdefault(dates, get_column_letter(start_col))

        start_col += 1

    wb.save(doc)
    return dates_dict

# working with the TSD files
def date_range_from_source(source_doc):
    wb = openpyxl.load_workbook(source_doc)
    sheet = wb.get_sheet_by_name('Sheet1')
    dates = []
    start_point = 'F2'
    finished_point = 'F' + str(sheet.max_row)
    print('Point 1')
    for row_of_cells in sheet[start_point:finished_point]:
        for cell in row_of_cells:
            if cell.value is None:
                pass
            elif type(cell.value) == datetime:
                dates.append(reformat_date(cell.value))
            else:
                dates.append(string_to_date(cell.value))
    wb.save(source_doc)
    return dates


def reformat_date(date_value):
    if date_value.day > 12:
        year = date_value.year
        month = date_value.month
        day = date_value.day
    else:
        year = date_value.year
        month = date_value.day
        day = date_value.month

    return date(year, month, day)


def string_to_date(date_string):

    temp = date_string.split(' ')
    temp = temp[0].split('/')
    return date(int(temp[2]), int(temp[1]), int(temp[0]))


# working with the finished file
def main_source(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.get_sheet_by_name('Register')
    dates = get_rev_dates(sheet)
    main_files = get_main_files(sheet)
    dates.sort()
    wb.save(path)
    return dates, main_files


def get_rev_dates(sheet):
    start_point = 'H1'
    finish_point = get_column_letter(sheet.max_column) + '1'
    dates = []
    for row_of_cells in sheet[start_point:finish_point]:
        for cell in row_of_cells:
            day = cell.value
            month = sheet.cell(row=cell.row + 1, column=column_index_from_string(cell.column)).value
            year = sheet.cell(row=cell.row + 2, column=column_index_from_string(cell.column)).value
            try:
                output = date(int('20'+str(year)), int(month), int(day))
            except TypeError:
                output = None
            except ValueError:
                output = None

            if output is not None:
                dates.append(output)

    return dates


def get_main_files(sheet):
    files = []
    max_rows = 'E' + str(sheet.max_row)
    start_point = 'E5'
    for row_cells in sheet[start_point:max_rows]:
        for cell in row_cells:
            if cell.value is None:
                break
            x = MainFile()
            x.build_main_file(cell.coordinate, sheet)
            files.append(x)
    return files


class MainFile:
    def __init__(self):
        self.drawing_number = None
        self.drawing_title = None
        self.revision = {}
        self.is_TSD = False
        self.is_in_progress = False
        self.is_error = False
        self.is_ignore = False

    def add_revision(self, rev, date_values):
        if rev is not None:
            self.revision.setdefault(rev, date(*date_values))

    def build_main_file(self, cell, sheet):
        # TODO this needs to be filled out.
        box = sheet.cell(coordinate=cell)
        row = box.row
        max_col = sheet.max_column
        start_col = 8
        self.is_TSD = sheet.cell(row=row, column=1).value
        self.is_in_progress = sheet.cell(row=row, column=2).value
        self.is_error = sheet.cell(row=row, column=3).value
        self.is_ignore = sheet.cell(row=row, column=4).value
        self.drawing_number = sheet.cell(row=row, column=5).value
        self.drawing_title = sheet.cell(row=row, column=6).value

        while start_col <= max_col:
            rev_date = []
            rev = sheet.cell(row=row, column=start_col).value
            rev_date.append(sheet.cell(row=3, column=start_col).value)
            rev_date.append(sheet.cell(row=2, column=start_col).value)
            rev_date.append(sheet.cell(row=1, column=start_col).value)
            self.add_revision(rev, rev_date)

            start_col += 1

    def build_tsd_file(self, row, sheet):
        self.is_TSD = True
        self.drawing_number = sheet.cell(row=row, column=2).value
        self.drawing_title = sheet.cell(row=row, column=3).value

        rev = sheet.cell(row=row, column=4).value
        rev_date = sheet.cell(row=row, column=6).value
        if type(rev_date) == datetime:

            rev_date = reformat_date(rev_date)
        elif rev_date is None:
            pass
        else:
            rev_date = reformat_date(string_to_date(rev_date))
        self.revision.setdefault(rev, rev_date)

    @staticmethod
    def reformat_date(date_value):
        if date_value.day > 12:
            year = date_value.year
            month = date_value.month
            day = date_value.day
        else:
            year = date_value.year
            month = date_value.day
            day = date_value.month

        return date(year, month, day)

    @staticmethod
    def string_to_date(date_string):

        temp = date_string.split(' ')
        temp = temp[0].split('/')
        return date(int(temp[2]), int(temp[1]), int(temp[0]))



    def __lt__(self, other):
        return self.drawing_number < other.drawing_number

    def __repr__(self):
        return '<Drawing No. {}>'.format(self.drawing_number)
