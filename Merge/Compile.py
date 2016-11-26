import openpyxl
import os
from pprint import pprint
from Merge.merge import find_file_type, newest_file
from pathlib import Path as p
__author__ = 'Desktop'


class PdfFile:
    def __init__(self, file):
        # file = p('')
        self.name = file.name
        self.path = str(file)
        self.version = file.lstat().st_ctime

    def __lt__(self, other):
        return self.name < other.name


class File:
    def __init__(self, sheet, row):
        self.row = row
        self.name = self.build_name(sheet)
        self.tsd = self.get_value(sheet, 1)
        self.in_progress = self.get_value(sheet, 2)
        self.in_error = self.get_value(sheet, 2)
        self.ignore = self.get_value(sheet, 2)
        self.path = ''

    def get_value(self, sheet, cell):
        if sheet.cell(row=self.row, column=cell).value is None:
            return False
        else:
            return True

    def build_name(self, sheet):
        drawing = sheet.cell(row=self.row, column=5).value
        rev = self.newest_rev(sheet)
        if rev is None:
            self.in_error = True

        name = '{}_{}.pdf'.format(drawing, rev)
        return name

    def newest_rev(self, sheet):
        column_count = sheet.max_column
        counter = 7
        revs = []
        ignore = ['Z', 'X', 'SS']
        output = ''
        while counter <= column_count:
            if sheet.cell(row=self.row, column=counter).value is not None:
                revs.append(sheet.cell(row=self.row, column=counter).value)
            counter += 1

        revs.sort()
        for rev in revs:
            if rev not in ignore:
                output = rev
                break
        return output

    def __repr__(self):
        return '<{} : row {}>'.format(self.name, self.row)

    def __lt__(self, other):
        return self.name < other.name


def find_pdf(root):

    for root, dirs, files in os.walk(root):

        for file in files:
            file = p(os.path.join(root, file))

            if find_file_type(file.suffix):
                yield file


def make_dict(pdf_list):
    keys = pdf_list.keys()
    output = {}
    for key in keys:
        output.setdefault(key, pdf_list[key].path)
    return output


def working_list(root):
    pdfs = (PdfFile(i) for i in find_pdf(root))
    pdfs = newest_file(pdfs)
    pdfs = make_dict(pdfs)
    return pdfs


def set_up_workbook(document):
    wb = openpyxl.load_workbook(document)
    sheet = wb.get_sheet_by_name("Register")
    return sheet


def get_files(document):
    output = []
    sheet = set_up_workbook(document)
    row = 5

    while row <= sheet.max_row:
        file = File(sheet, row)
        output.append(file)
        row += 1
    return output


def safe_keys(pdfs):
    output = {}
    keys = pdfs.keys()
    for key in list(keys):
        temp = key.split('.')
        output.setdefault('{}.{}'.format(temp[0], temp[1].capitalize()), key)
        output.setdefault('{}.{}'.format(temp[0], temp[1].lower()), key)
    return output

def find_paths(files, pdfs):
    keys = safe_keys(pdfs)
    for file in files:
        if file.name in keys.keys():
            file.path = pdfs[keys[file.name]]
        else:
            file.in_error = True

def error_count(files):
    x = 0
    for file in files:
        if file.in_error:
            x += 1
    print("Number of files with errors: {}".format(x))


def run(document, root, destination):
    pdfs = working_list(root)

    files = get_files(document)

    find_paths(files, pdfs)

    pprint(pdfs)
    files.sort()
    for file in files:
        pprint(file)

    error_count(files)



